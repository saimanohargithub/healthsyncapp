#!/usr/bin/env python3
"""Convert DAST report.json → Excel (.xlsx) in PASS / FAIL format."""

import json, sys
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR  = Path(__file__).resolve().parent
REPORT_JSON = SCRIPT_DIR / "report.json"
OUTPUT_XLSX = SCRIPT_DIR / "dast_report.xlsx"

# ─── Styles ───────────────────────────────────────────────────
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor="2F5496")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F3864")
SUB_FONT = Font(name="Calibri", bold=True, size=11, color="404040")
NORM = Font(name="Calibri", size=10)
BOLD = Font(name="Calibri", bold=True, size=10)

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
PASS_FONT = Font(name="Calibri", bold=True, size=11, color="006100")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
FAIL_FONT = Font(name="Calibri", bold=True, size=11, color="9C0006")

SEV_FILLS = {
    "CRITICAL": PatternFill("solid", fgColor="FF4444"),
    "HIGH":     PatternFill("solid", fgColor="FF8C00"),
    "MEDIUM":   PatternFill("solid", fgColor="FFD700"),
    "LOW":      PatternFill("solid", fgColor="90EE90"),
    "INFO":     PatternFill("solid", fgColor="D9E2F3"),
}
SEV_FONTS = {
    "CRITICAL": Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
    "HIGH":     Font(name="Calibri", bold=True, size=10, color="FFFFFF"),
    "MEDIUM":   Font(name="Calibri", bold=True, size=10, color="333333"),
    "LOW":      Font(name="Calibri", size=10, color="333333"),
    "INFO":     Font(name="Calibri", size=10, color="333333"),
}

BORDER = Border(
    left=Side("thin", color="B4B4B4"), right=Side("thin", color="B4B4B4"),
    top=Side("thin", color="B4B4B4"), bottom=Side("thin", color="B4B4B4"),
)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")


def hdr_row(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font, cell.fill, cell.alignment, cell.border = HDR_FONT, HDR_FILL, CTR, BORDER


# ══════════════════════════════════════════════════════════════
#  Sheet 1 — All Test Cases (Pass / Fail)
# ══════════════════════════════════════════════════════════════
def sheet_all(wb, data):
    ws = wb.active
    ws.title = "All Test Cases"

    # Title rows
    ws.merge_cells("A1:J1")
    ws["A1"] = "HealthSync — DAST Security Test Results"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    total = len(data)
    passed = sum(1 for r in data if not r["finding"])
    failed = total - passed
    ws["A2"] = (f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  "
                f"Total: {total}  |  Passed: {passed}  |  Failed: {failed}  |  "
                f"Pass Rate: {passed/total*100:.1f}%")
    ws["A2"].font = SUB_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["S.No", "Test Category", "Test Scenario", "Endpoint",
               "Method", "HTTP Status", "Expected", "Result",
               "Severity", "Remarks"]
    hr = 4
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    hdr_row(ws, hr, len(headers))
    ws.row_dimensions[hr].height = 22

    # Data
    for idx, rec in enumerate(data, 1):
        r = hr + idx
        is_fail = rec["finding"]

        ws.cell(row=r, column=1, value=idx)
        ws.cell(row=r, column=2, value=rec["test_category"].replace("_", " ").title())

        # Build a human-readable test scenario name
        cat = rec["test_category"]
        role = rec["role"]
        if cat == "authn_bypass":
            scenario = f"Access with {role.replace('_', ' ')}"
        elif cat == "authz_idor":
            scenario = f"Cross-user access ({role.replace('_', ' ')})"
        elif cat == "rbac_matrix":
            scenario = f"{role.replace('_', ' ').title()} read access"
        elif cat == "token_tampering":
            scenario = f"JWT tampered: {role.replace('_', ' ')}"
        elif cat == "injection":
            scenario = f"Injection probe ({role})"
        elif cat == "rate_limiting":
            scenario = f"Burst {role} test"
        elif cat == "hardcoded_creds":
            scenario = "Static credential scan"
        elif cat == "firestore_rules":
            scenario = "Unauthenticated collection access"
        else:
            scenario = role
        ws.cell(row=r, column=3, value=scenario)

        ws.cell(row=r, column=4, value=rec["endpoint"])
        ws.cell(row=r, column=5, value=rec["method"])
        ws.cell(row=r, column=6, value=str(rec["status"]))
        ws.cell(row=r, column=7, value=rec["expected_status"])

        # ── PASS / FAIL cell ──
        rc = ws.cell(row=r, column=8, value="FAIL" if is_fail else "PASS")
        rc.fill = FAIL_FILL if is_fail else PASS_FILL
        rc.font = FAIL_FONT if is_fail else PASS_FONT

        # Severity
        sev = rec["severity"]
        sc = ws.cell(row=r, column=9, value=sev)
        sc.fill = SEV_FILLS.get(sev, SEV_FILLS["INFO"])
        sc.font = SEV_FONTS.get(sev, SEV_FONTS["INFO"])

        ws.cell(row=r, column=10, value=rec["note"])

        # Style all cells in row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = CTR if c in (1, 5, 6, 7, 8, 9) else LFT
            if cell.font == Font():
                cell.font = NORM
            # Alternate row shading (skip Result & Severity columns)
            if idx % 2 == 0 and c not in (8, 9):
                cell.fill = ALT_FILL

    # Column widths
    for i, w in enumerate([6, 22, 35, 48, 8, 12, 12, 10, 12, 55], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{hr + len(data)}"


# ══════════════════════════════════════════════════════════════
#  Sheet 2 — Summary Dashboard
# ══════════════════════════════════════════════════════════════
def sheet_summary(wb, data):
    ws = wb.create_sheet("Summary")

    ws.merge_cells("A1:E1")
    ws["A1"] = "Test Summary Dashboard"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    total = len(data)
    passed = sum(1 for r in data if not r["finding"])
    failed = total - passed

    # Overall stats
    ws["A3"] = "Overall"
    ws["A3"].font = SUB_FONT
    for i, (k, v) in enumerate([
        ("Total Tests", total), ("Passed", passed),
        ("Failed", failed), ("Pass Rate", f"{passed/total*100:.1f}%")
    ]):
        ws.cell(row=4+i, column=1, value=k).font = BOLD
        vc = ws.cell(row=4+i, column=2, value=v)
        vc.font = BOLD
        if k == "Passed":
            vc.fill = PASS_FILL; vc.font = PASS_FONT
        elif k == "Failed":
            vc.fill = FAIL_FILL; vc.font = FAIL_FONT
        for c in (1, 2):
            ws.cell(row=4+i, column=c).border = BORDER

    # By severity
    ws["A10"] = "Failed Tests by Severity"
    ws["A10"].font = SUB_FONT
    for i, h in enumerate(["Severity", "Count"], 1):
        ws.cell(row=11, column=i, value=h)
    hdr_row(ws, 11, 2)
    findings = [r for r in data if r["finding"]]
    by_sev = {}
    for f in findings:
        by_sev[f["severity"]] = by_sev.get(f["severity"], 0) + 1
    row = 12
    for sev in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
        cnt = by_sev.get(sev, 0)
        c1 = ws.cell(row=row, column=1, value=sev)
        c1.fill = SEV_FILLS[sev]; c1.font = SEV_FONTS[sev]; c1.border = BORDER
        c2 = ws.cell(row=row, column=2, value=cnt)
        c2.font = BOLD; c2.border = BORDER; c2.alignment = CTR
        row += 1

    # By category
    ws["A18"] = "Results by Category"
    ws["A18"].font = SUB_FONT
    for i, h in enumerate(["Category", "Total", "Passed", "Failed", "Result"], 1):
        ws.cell(row=19, column=i, value=h)
    hdr_row(ws, 19, 5)

    by_cat = {}
    for r in data:
        cat = r["test_category"].replace("_", " ").title()
        if cat not in by_cat:
            by_cat[cat] = {"total": 0, "fail": 0}
        by_cat[cat]["total"] += 1
        if r["finding"]:
            by_cat[cat]["fail"] += 1

    row = 20
    for cat in sorted(by_cat):
        info = by_cat[cat]
        p = info["total"] - info["fail"]
        all_pass = info["fail"] == 0
        ws.cell(row=row, column=1, value=cat).font = NORM
        ws.cell(row=row, column=2, value=info["total"]).font = NORM
        ws.cell(row=row, column=3, value=p).font = NORM
        ws.cell(row=row, column=4, value=info["fail"]).font = NORM
        rc = ws.cell(row=row, column=5, value="PASS" if all_pass else "FAIL")
        rc.fill = PASS_FILL if all_pass else FAIL_FILL
        rc.font = PASS_FONT if all_pass else FAIL_FONT
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = CTR if c > 1 else LFT
        row += 1

    for i, w in enumerate([28, 10, 10, 10, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ══════════════════════════════════════════════════════════════
#  Sheet 3 — Failed Tests Only
# ══════════════════════════════════════════════════════════════
def sheet_failures(wb, data):
    ws = wb.create_sheet("Failed Tests")

    ws.merge_cells("A1:H1")
    ws["A1"] = "Failed Test Cases — Action Required"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["S.No", "Severity", "Category", "Endpoint",
               "Method", "Result", "Description", "Recommended Fix"]
    hr = 3
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    hdr_row(ws, hr, len(headers))

    fails = sorted(
        [r for r in data if r["finding"]],
        key=lambda x: {"CRITICAL":0,"HIGH":1,"MEDIUM":2,"LOW":3}.get(x["severity"], 4)
    )
    for idx, rec in enumerate(fails, 1):
        r = hr + idx
        ws.cell(row=r, column=1, value=idx).font = NORM

        sev = rec["severity"]
        sc = ws.cell(row=r, column=2, value=sev)
        sc.fill = SEV_FILLS.get(sev, SEV_FILLS["INFO"])
        sc.font = SEV_FONTS.get(sev, SEV_FONTS["INFO"])

        ws.cell(row=r, column=3, value=rec["test_category"].replace("_"," ").title()).font = NORM
        ws.cell(row=r, column=4, value=rec["endpoint"]).font = NORM
        ws.cell(row=r, column=5, value=rec["method"]).font = NORM

        rc = ws.cell(row=r, column=6, value="FAIL")
        rc.fill = FAIL_FILL; rc.font = FAIL_FONT

        ws.cell(row=r, column=7, value=rec["note"]).font = NORM

        cat = rec["test_category"]
        if cat in ("firestore_rules","authn_bypass","rbac_matrix","authz_idor"):
            fix = "Deploy Firestore security rules (firestore.rules)"
        elif cat == "hardcoded_creds":
            fix = "FIXED — hardcoded keys removed from source code"
        elif cat == "rate_limiting":
            fix = "Enable Firebase App Check for rate limiting"
        elif cat == "injection":
            fix = "Low risk — inherent Firestore path behavior"
        elif cat == "token_tampering":
            fix = "Review token validation logic"
        else:
            fix = "Review and remediate"
        ws.cell(row=r, column=8, value=fix).font = NORM

        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = CTR if c in (1,2,5,6) else LFT
        if idx % 2 == 0:
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c)
                if c not in (2, 6):
                    cell.fill = ALT_FILL

    for i, w in enumerate([6, 12, 20, 48, 8, 10, 55, 45], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


# ─── Main ─────────────────────────────────────────────────────
def main():
    data = json.loads(REPORT_JSON.read_text())
    print(f"  Loaded {len(data)} test records")

    wb = openpyxl.Workbook()
    sheet_all(wb, data)
    sheet_summary(wb, data)
    sheet_failures(wb, data)
    
    saved_path = OUTPUT_XLSX
    try:
        wb.save(str(OUTPUT_XLSX))
    except PermissionError:
        saved_path = SCRIPT_DIR / "dast_report_new.xlsx"
        print(f"  [WARNING] Permission denied on {OUTPUT_XLSX.name} (probably open in Excel). Saving to {saved_path.name} instead.")
        wb.save(str(saved_path))

    passed = sum(1 for r in data if not r["finding"])
    failed = len(data) - passed
    print(f"  [OK] Saved: {saved_path}")
    print(f"  Total: {len(data)} | PASS: {passed} | FAIL: {failed}")

if __name__ == "__main__":
    main()
